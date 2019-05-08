# -*- coding: utf-8 -*-
from time import *
from openpyxl import load_workbook
import xlsxwriter
import sys
sys.path.append('..')
from TestModel import Functions, MyUnitTest
from TestElements import Common_HomeScreen, Common_SearchScreen, Common_NewAccountDetailsScreen, \
    Common_AccountSummaryScreen, Common_StartAQuoteScreen, GL_CoveragePartsScreen, Common_BasicPolicyInfoScreen, \
    Common_LocationsScreen, GL_CGLScreen, Common_LossHistoryScreen, Common_PolicyReviewScreen, \
    Common_QuoteSummaryScreen, Common_ReferToUnderwriterScreen, Common_PCScreen, Common_PayAndIssueScreen

class CreateGLIndicativeIssued(MyUnitTest.MyTest):

    def test_new_gl_indicative_Issued(self):

        # Create screenshot excel
        book = xlsxwriter.Workbook('..\MIG_Portal_Automation\TestReports\Create_GL_Indicative_Issued.xlsx')

        # Load test data
        wb = load_workbook('..\MIG_Portal_Automation\TestData/test_Create_GL_Indicative_Issued.xlsx')
        ws = wb['GLIssued']
        AccountNumber = ws.cell(2, 1).value
        ClassCode = ws.cell(2, 2).value
        AnnualBasis = ws.cell(2, 3).value

        # Login Portal
        Functions.login_portal(self.driver)

        # Home page
        homeScreen = Common_HomeScreen.Home(self.driver)
        homeScreen.click_search_icon()

        # Search Screen
        searchScreen = Common_SearchScreen.Search(self.driver)
        searchScreen.input_searchkeyword(AccountNumber)
        Functions.get_screenshot(self.driver, book, "SearchAccount")
        searchScreen.click_view_button()

        # Account Summary Screen
        accountSummary = Common_AccountSummaryScreen.AccountSummary(self.driver)
        Functions.get_screenshot(self.driver, book, "AccountSummary")
        accountSummary.click_startquote()

        # Start A Quote Screen
        startQuote = Common_StartAQuoteScreen.StartAQuote(self.driver)
        startQuote.click_indicative()
        Functions.get_screenshot(self.driver, book, "StartAQuote")
        startQuote.click_generalliability()
        startQuote.click_next_button()

        # GL Coverage Parts Screen
        coverageParts = GL_CoveragePartsScreen.CoverageParts(self.driver)
        coverageParts.click_commercialgeneralliability()
        Functions.get_screenshot(self.driver, book, "CoverageParts")
        coverageParts.click_epli()
        coverageParts.click_cyberliability()
        coverageParts.click_next_button()

        # Policy Info Screen
        policyInfo = Common_BasicPolicyInfoScreen.BasicPolicyInfo(self.driver)
        submissionnumber = policyInfo.get_submission_number()
        print(submissionnumber)
        Functions.get_screenshot(self.driver, book, "PolicyInfo")
        policyInfo.click_next_button()

        # Locations Screen
        locations = Common_LocationsScreen.Locations(self.driver)
        Functions.get_screenshot(self.driver, book, "Locations")
        locations.click_next_button()

        # CGL Screen
        cgl = GL_CGLScreen.CGL(self.driver)
        cgl.click_classcodes_tab()
        cgl.input_classcode(ClassCode)
        Functions.get_screenshot(self.driver, book, "ClassCodes")
        cgl.input_annualbasis(AnnualBasis)
        cgl.click_next_button()

        # Loss History
        losshistory = Common_LossHistoryScreen.LossHistory(self.driver)
        losshistory.select_lengthofpriorlosshistory("Unknown")
        losshistory.select_priorlosshistory("Unknown")
        Functions.get_screenshot(self.driver, book, "Loss History")
        losshistory.click_next_button()

        # Policy Review
        policyreview = Common_PolicyReviewScreen.PolicyReview(self.driver)
        Functions.get_screenshot(self.driver, book, "PolicyReview")
        policyreview.click_quote_button()

        # Quote Summary
        quotesummary = Common_QuoteSummaryScreen.QuoteSummary(self.driver)
        print(quotesummary.get_quote_status())
        # self.assertEqual(quotesummary.get_quote_status(), 'QUOTED')

        # Agent Notes
        refertounderwriter = Common_ReferToUnderwriterScreen.ReferToUnderwriter(self.driver)
        #refertounderwriter.input_note("amy_notes")
        #Functions.get_screenshot(self.driver, book, "AddNote")
        refertounderwriter.click_cancel()

        # Make issueable quote
        quotesummary.click_makeissuablequote()
        Functions.get_screenshot(self.driver, book, "QuoteSummary")
        quotesummary.click_gotoerror()
        sleep(2)
        quotesummary.click_folderror()
        #Functions.page_scroll_down(self.driver)
        Functions.get_screenshot(self.driver, book, "SupplementalInfo")
        cgl.answer_all_questions()
        cgl.click_next_button()
        losshistory.click_next_button()
        policyreview.click_quote_button()
        refertounderwriter.input_note("amy_notes")
        Functions.get_screenshot(self.driver, book, "AddNote")
        refertounderwriter.click_submit()

        # Policy Center
        pc = Common_PCScreen.PC(self.driver)
        Functions.login_pc(self.driver)
        pc.search_submission(submissionnumber)
        Functions.get_screenshot(self.driver, book, "PC")
        pc.click_riskanalysis()
        pc.checkon_uwissues()
        pc.click_approve()
        pc.click_ok()
        pc.uncheck_hide()
        pc.click_next()

        # Return to portal to issue
        Functions.login_portal(self.driver)
        homeScreen.click_search_icon()
        searchScreen.input_searchkeyword(submissionnumber)
        searchScreen.click_view_button()
        Functions.get_screenshot(self.driver, book, "PayandIssue")
        payandissue = Common_PayAndIssueScreen.PayAndIssue(self.driver)
        payandissue.click_payandissue_button()
        Functions.get_screenshot(self.driver, book, "IssueSucsess")

        book.close()







