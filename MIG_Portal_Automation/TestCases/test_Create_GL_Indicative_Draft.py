# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import xlsxwriter
import sys
sys.path.append('..')
from TestModel import Functions, MyUnitTest
from TestElements import Common_HomeScreen, Common_SearchScreen, Common_NewAccountDetailsScreen, \
    Common_AccountSummaryScreen, Common_StartAQuoteScreen, GL_CoveragePartsScreen

class CreateGLIndicativeDraft(MyUnitTest.MyTest):

    def test_new_gl_draft(self):

        # Create screenshot excel
        book = xlsxwriter.Workbook('..\MIG_Portal_Automation\TestReports\Create_GL_Indicative_Draft.xlsx')

        # Load test data
        wb = load_workbook('..\MIG_Portal_Automation\TestData/test_Create_GL_Indicative_Draft.xlsx')
        ws = wb['GLDraft']
        AccountNumber = ws.cell(2, 1).value

        # Login Portal
        Functions.login_portal(self.driver)
       
        # Home page
        homeScreen = Common_HomeScreen.Home(self.driver)
        homeScreen.click_search_icon()

        # Search Screen
        searchScreen = Common_SearchScreen.Search(self.driver)
        searchScreen.input_searchkeyword(AccountNumber)
        Functions.get_screenshot(self.driver, book, "SearchAccount")
        searchScreen.click_view_button()

        # Account Summary Screen
        accountSummary = Common_AccountSummaryScreen.AccountSummary(self.driver)
        Functions.get_screenshot(self.driver, book, "AccountSummary")
        accountSummary.click_startquote()

        # Start A Quote Screen
        startQuote = Common_StartAQuoteScreen.StartAQuote(self.driver)
        startQuote.click_indicative()
        Functions.get_screenshot(self.driver, book, "StartAQuote")
        startQuote.click_generalliability()
        startQuote.click_next_button()

        # GL Coverage Parts Screen
        coverageParts = GL_CoveragePartsScreen.CoverageParts(self.driver)
        quotestatus = coverageParts.get_quote_status()
        self.assertEqual(quotestatus, 'DRAFT')
        print(coverageParts.get_quote_status())
        Functions.get_screenshot(self.driver, book, "CoverageParts")

        book.close()








