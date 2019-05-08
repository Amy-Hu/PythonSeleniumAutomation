# -*- coding: utf-8 -*-
from datetime import datetime
from openpyxl import load_workbook
import xlsxwriter
import sys
sys.path.append('..')
from TestModel import Functions, MyUnitTest
from TestElements import Common_HomeScreen, Common_SearchScreen, Common_NewAccountDetailsScreen, Common_AccountSummaryScreen

class CreateNewAccount(MyUnitTest.MyTest):

    def test_new_account(self):

        # Create screenshot excel
        book = xlsxwriter.Workbook('..\MIG_Portal_Automation\TestReports\Create_NewAccount.xlsx')

        # Load test data
        wb = load_workbook('..\MIG_Portal_Automation\TestData/test_Create_New_Account.xlsx')
        ws = wb['NewAccount']
        Company = ws.cell(2, 1).value + datetime.now().strftime('%Y%m%d%H%M%S')
        Address1 = ws.cell(2, 2).value
        ZipCode = ws.cell(2, 3).value
        City = ws.cell(2, 4).value
        County = ws.cell(2, 5).value
        State = ws.cell(2, 6).value
       
        # Login Portal
        Functions.login_portal(self.driver)
       
        # Home Screen
        homeScreen = Common_HomeScreen.Home(self.driver)
        homeScreen.click_search_icon()

        # Search Screen
        searchScreen = Common_SearchScreen.Search(self.driver)
        searchScreen.click_createanewone_button()
        
        # New Account Details Screen
        newAccount = Common_NewAccountDetailsScreen.NewAccount(self.driver)
        newAccount.input_company(Company)
        Functions.get_screenshot(self.driver, book, "NewAccount")
        newAccount.click_currentcustomer_no()
        newAccount.input_address(Address1)
        newAccount.input_postalcode(ZipCode)
        newAccount.input_city(City)
        newAccount.select_state(State)
        newAccount.input_county(County)
        newAccount.select_addresstype()
        newAccount.input_officephone("201-555-1234")
        newAccount.select_contactmethod()
        newAccount.click_next_button()

        # Account Summary Screen
        accountSummary = Common_AccountSummaryScreen.AccountSummary(self.driver)
        accountStatus = accountSummary.get_accountstatus()
        accountnumber = accountSummary.get_accountnumber()
        self.assertEqual(accountStatus, 'Pending')
        print(accountnumber)
        Functions.get_screenshot(self.driver, book, "AccountSummary")
        book.close()






