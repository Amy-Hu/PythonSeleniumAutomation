# -*- coding: utf-8 -*-
from selenium import webdriver
from time import *
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import xlsxwriter
from pathlib import Path
from openpyxl import load_workbook
import os, sys

def get_screenshot(driver, book, sheet):
    i = 1
    scrpath = "..\MIG_Portal_Automation\TestReports\Screenshots"
    capturename = '\\Screenshot' + str(i) + '.png'
    wholepath = scrpath + capturename
    if Path(scrpath).is_dir():
        pass
    else:
        Path(scrpath).mkdir()
    while Path(wholepath).exists():
        i += 1
        capturename = '\\Screenshot' + str(i) + '.png'
        wholepath = scrpath + capturename
    driver.get_screenshot_as_file(wholepath)
    book.add_worksheet(sheet).insert_image('A1',wholepath)

def check_alert_display(driver):
    try:
        alert = driver.switch_to.alert
        return alert
    except:
        return False

def wait_element_visible(driver, byType, value):
    try:
        WebDriverWait(driver, 5).until(EC.visibility_of_element_located((byType, value)))
        return True
    except:
        return False

def wait_element_not_visible(driver, byType, value):
    try:
        WebDriverWait(driver, 30).until_not(EC.visibility_of_element_located((byType, value)))
        return True
    except:
        return False

def page_scroll_down(driver):
    js = "var q=document.documentElement.scrollTop=80"
    driver.execute_script(js)

def login_portal(driver):
    wb = load_workbook('..\MIG_Portal_Automation\TestData/Common_Login_Portal.xlsx')
    ws = wb['LoginInformation']
    if ws.cell(2, 1).value == "Pre Prod":
        url = "https://agent.pre-prod.motoristsinsurance.com/"
    if ws.cell(2, 1).value == "QA Monthly":
        url = "https://agent.monthly.qa.motoristsinsurance.com/"
    if ws.cell(2, 1).value == "QA Daily":
        url = "https://agent.qa.motoristsinsurance.com/"
    LoginUser = ws.cell(2, 2).value
    Password = ws.cell(2, 3).value
    driver.get(url)
    alert = check_alert_display(driver)
    if alert:
        alert.accept()
    if wait_element_visible(driver, By.NAME, "username"):
        WebDriverWait(driver, 15, 0.5).until(EC.presence_of_element_located((By.NAME, "username"))).send_keys(LoginUser)
        WebDriverWait(driver, 15, 0.5).until(EC.element_to_be_clickable((By.ID, "okta-signin-submit"))).click()
        WebDriverWait(driver, 15, 0.5).until(EC.presence_of_element_located((By.NAME, "password"))).send_keys(Password)
        WebDriverWait(driver, 15, 0.5).until(EC.element_to_be_clickable((By.XPATH, "//input[@type='submit']"))).click()

def login_pc(driver):
    wb = load_workbook('..\MIG_Portal_Automation\TestData/Common_Login_PC.xlsx')
    ws = wb['LoginInformation']
    if ws.cell(2, 1).value == "Pre Prod":
        url = "http://policy.pre-prod.test.mig.corp/pc/PolicyCenter.do"
    if ws.cell(2, 1).value == "QA Monthly":
        url = "http://qaiab.monthly.test.mig.corp:8180/pc/PolicyCenter.do"
    if ws.cell(2, 1).value == "QA Daily":
        url = "http://qaiab.test.mig.corp:8180/pc/PolicyCenter.do"
    LoginUser = ws.cell(2, 2).value
    Password = ws.cell(2, 3).value
    driver.get(url)
    WebDriverWait(driver, 15, 0.5).until(EC.presence_of_element_located((By.XPATH, "//input[@id='Login:LoginScreen:LoginDV:username-inputEl']"))).send_keys(LoginUser)
    WebDriverWait(driver, 15, 0.5).until(EC.presence_of_element_located((By.XPATH, "//input[@id='Login:LoginScreen:LoginDV:password-inputEl']"))).send_keys(Password)
    WebDriverWait(driver, 15, 0.5).until(EC.element_to_be_clickable((By.XPATH, "//span[@id='Login:LoginScreen:LoginDV:submit-btnInnerEl']"))).click()

